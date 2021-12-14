import time

import openpyxl
from geopy.geocoders import ArcGIS
import json
import time

def get_data():
    try:
        book = openpyxl.open("starbucks.xlsx", read_only=True)

        sheet = book.active
        for i in range(2, sheet.max_row+1): #sheet.max_row+1
            adress = sheet[i][1].value  # [row][column]
            name = sheet[i][0].value
            print(f"Сохраняю место с именем {name} {i}/{sheet.max_row}")
            try:
                response = ArcGIS(user_agent='Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36').geocode(adress)
                latitude = round(response.latitude,3)  # широта
                longitude = round(response.longitude,3)  # долгота
            except:
                latitude = 'Адрес не найден'
                longitude = 'Адрес не найден'

            adress_part = adress.replace(',', '').replace('.', ' ').split()
            try:
                city = list(set(adress_part) & set(cities))[0]
            except Exception as ex:
                city = 'Город не найден'
            places = {
                "coords": [latitude, longitude],
                "name": name,
                "address": adress
            }
            write_json(city, places)
    except Exception as ex:
        print(ex)
        print('Файл starbucks.xlsx не найден')

cities = ['Москва', 'Московская область', 'Санкт-Петербург', 'Казань',
    'Самара', 'Екатеринбург', 'Тюмень',
    'Ярославль',
    'Краснодар', 'Сочи']

cities_json = {
    'Москва': 'moscow', 'Московская область': 'moscow_obl', 'Казань': 'kazan',
    'Самара': 'samara', 'Екатеринбург': 'ekaterinburg', 'Тюмень': 'tymen',
    'Ярославль': 'yaroslavl',
    'Краснодар': 'krasnodar', 'Сочи': 'sochi', 'Санкт-Петербург' : 'sankt_peterburg',
    'Город не найден': 'empty'}


# Сохраняем в формат JSON и открываем файл
def write_json(city, places):
    with open('cities/'+cities_json[city]+'.json', "a", encoding="utf-8") as file:
        json.dump(places, file, ensure_ascii=False)
        file.write(',\n')


def main():
    get_data()
    time.sleep(5)


if __name__== "__main__":
    main()