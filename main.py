import openpyxl
from geopy.geocoders import ArcGIS
import json

def get_data():

    book = openpyxl.open("starbucks.xlsx", read_only=True)
    sheet = book.active
    for i in range(130,sheet.max_row+1): #sheet.max_row+1
        #adress = sheet[i][1].value +','+ sheet[i][2].value #[row][column]
        adress = sheet[i][1].value  # [row][column]
        name = sheet[i][0].value

        print(f"Сохраняю место с именем {name} {i}/{sheet.max_row}")
        try:
            #response = Nominatim(user_agent='Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36').geocode(adress)
            response = ArcGIS(user_agent='Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36').geocode(adress)
            latitude = response.latitude  # широта
            longitude = response.longitude  # долгота
        except:
            latitude = 'Адрес не найден'
            longitude = 'Адрес не найден'

        adress_part = adress.replace(',', '').replace('.', ' ').split()
        try:
            city = list(set(adress_part) & set(cities))[0]
        except:
            city = 'Город не найден'
        places = []
        places.append({
            "coords": [latitude, longitude],
            "name": name,
            "adress": adress
        })
        write_json(city, places)

cities = ['Москва', 'Московская область', 'Санкт-Петербург', 'Казань',
    'Самара', 'Екатеринбург', 'Тюмень',
    'Ярославль',
    'Краснодар', 'Сочи']

cities_json = {
    'Москва': 'moscow.json', 'Московская область': 'moscow_obl.json', 'Казань': 'Kazan.json',
    'Самара': 'samara.json', 'Екатеринбург': 'ekaterinburg.json', 'Тюмень': 'tymen.json',
    'Ярославль': 'yaroslavl.json',
    'Краснодар': 'krasnodar.json', 'Сочи': 'Sochi.json', 'Санкт-Петербург' : 'Sankt_Peterburg.json',
    'Город не найден': 'empty.json'}


# Сохраняем в формат JSON и открываем файл
def write_json(city, places):
    with open(cities_json[city], "a", encoding="utf-8") as file:
        json.dump(places, file, indent=4, ensure_ascii=False)
        #file.write('\n')


def main():
    get_data()
    #write_json()

if __name__== "__main__":
    main()